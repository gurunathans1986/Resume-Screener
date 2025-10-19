import os
import re
import json
import argparse
import docx2txt
import PyPDF2
import time
import pandas as pd
import google.generativeai as genai
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from dotenv import load_dotenv


load_dotenv()  # Load your GEMINI_API_KEY from .env
# ---------------- Helper Functions ---------------- #

genai.configure(api_key=os.getenv("GEMINI_API_KEY"))

def clean_text(text: str) -> str:
    """Basic text cleanup."""
    text = text.strip()
    text = re.sub(r'\s+', ' ', text)
    return text

def extract_text_from_pdf(file_path: str) -> str:
    """Extract text from a PDF file."""
    text = ""
    with open(file_path, "rb") as f:
        reader = PyPDF2.PdfReader(f)
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text
    return text

def extract_text_from_docx(file_path: str) -> str:
    """Extract text from a Word file."""
    return docx2txt.process(file_path)

def read_resume(file_path: str) -> str:
    """Read resume text from PDF, DOCX, or TXT."""
    ext = os.path.splitext(file_path)[-1].lower()
    if ext == ".pdf":
        return extract_text_from_pdf(file_path)
    elif ext == ".docx":
        return extract_text_from_docx(file_path)
    elif ext == ".txt":
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()
    else:
        raise ValueError(f"Unsupported file format: {ext}")

def score_resume_with_gemini(jd_text: str, resume_text: str, model: str = "gemini-2.5-flash") -> dict:
    """Send JD + Resume to Gemini, get structured JSON back (strict JSON)."""
    prompt = f"""
    You are an expert recruiter. Compare the following Job Description (JD) and Resume.

    ##Output
    Return ONLY a valid JSON object in this exact format, nothing else:
    
    {{
      "score": <0-100 overall suitability>,
      "tech_stack_match": "<text>",
      "years_of_experience_match": "<text>",
      "domain_match": "<text>",
      "leadership": "<text>",
      "red_flags": "<text>",
      "reasoning": "<3-4 lines overall summary>"
    }}
    ##Input
    JD:
    {jd_text}

    Resume:
    {resume_text}
    """

    response = genai.GenerativeModel(model).generate_content(prompt)
    raw_text = (response.text or "").strip()

    # Cleanup accidental code fences like ```json ... ```
    if raw_text.startswith("```"):
        raw_text = raw_text.strip("`")
        if raw_text.lower().startswith("json"):
            raw_text = raw_text[4:].strip()

    try:
        return json.loads(raw_text)
    except Exception:
        print("⚠️ JSON parsing failed. Raw response was:\n", raw_text)
        return {
            "score": 0,
            "tech_stack_match": "Parsing failed",
            "years_of_experience_match": "Parsing failed",
            "domain_match": "Parsing failed",
            "leadership": "Parsing failed",
            "red_flags": "Parsing failed",
            "reasoning": "Parsing failed"
        }

def apply_excel_styling(file_path: str):
    """Apply bold headers, Yes/No coloring, text wrapping, and auto column width."""
    wb = load_workbook(file_path)
    ws = wb.active

    # Bold headers
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Map headers to column indices
    header_map = {cell.value: cell.column for cell in ws[1]}

    yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
    no_fill  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red

    yes_no_cols = ["Shortlisted", "Sent Codility Test"]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_name in yes_no_cols:
            if col_name in header_map:
                col_idx = header_map[col_name] - 1  # zero-based index for row[]
                cell = row[col_idx]
                if cell.value == "Yes":
                    cell.fill = yes_fill
                elif cell.value == "No":
                    cell.fill = no_fill

    # Wrap text in Reasoning column
    if "Reasoning" in header_map:
        col_idx = header_map["Reasoning"] - 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row[col_idx].alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-fit column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(file_path)


# ---------------- Main Script ---------------- #

def evaluate_resumes(date_str: str, threshold: int = 70):
    job_desc_dir = "job_descriptions"
    resume_dir = "resumes"
    results_dir = "results"
    master_file = os.path.join(results_dir, "master_results.xlsx")

    os.makedirs(results_dir, exist_ok=True)

    all_results = []
    analysis_run_date = datetime.today().strftime("%Y-%m-%d")

    # Loop over all roles (JD files)
    for jd_file in os.listdir(job_desc_dir):
        role_name, ext = os.path.splitext(jd_file)
        if ext.lower() != ".txt":
            continue

        # Read JD
        with open(os.path.join(job_desc_dir, jd_file), "r", encoding="utf-8") as f:
            jd_text = clean_text(f.read())

        # Resumes path
        role_resume_dir = os.path.join(resume_dir, role_name, date_str)
        if not os.path.exists(role_resume_dir):
            print(f"No resumes found for role {role_name} on {date_str}. Skipping.")
            continue

        role_results = []

        for resume_file in os.listdir(role_resume_dir):
            resume_path = os.path.join(role_resume_dir, resume_file)
            try:
                resume_text = clean_text(read_resume(resume_path))
                evaluation = score_resume_with_gemini(jd_text, resume_text)

                score = evaluation.get("score", 0)
                shortlisted = "Yes" if score >= threshold else "No"

                row = {
                    "Date": date_str,  # batch date (from arg)
                    "Analysis Run Date": analysis_run_date,  # today's date
                    "Role": role_name,
                    "Resume Name": resume_file,
                    "Candidate Name": os.path.splitext(resume_file)[0],
                    "Score": score,
                    "Shortlisted": shortlisted,
                    "Tech Stack Match": evaluation.get("tech_stack_match", ""),
                    "Years of Exp Match": evaluation.get("years_of_experience_match", ""),
                    "Domain Match": evaluation.get("domain_match", ""),
                    "Leadership": evaluation.get("leadership", ""),
                    "Red Flags": evaluation.get("red_flags", ""),
                    "Reasoning": evaluation.get("reasoning", ""),
                    "Sent Codility Test": ""  # fill manually later
                }

                role_results.append(row)
                all_results.append(row)
                #time.sleep(35)

            except Exception as e:
                print(f"Error processing {resume_file} for role {role_name}: {e}")

        # Save per-role results (Excel)
        if role_results:
            role_df = pd.DataFrame(role_results)
            role_output_dir = os.path.join(results_dir, role_name)
            os.makedirs(role_output_dir, exist_ok=True)
            role_file = os.path.join(role_output_dir, f"{date_str}_results.xlsx")
            with pd.ExcelWriter(role_file, engine="openpyxl") as writer:
                role_df.to_excel(writer, index=False)
            apply_excel_styling(role_file)
            print(f"Saved results for {role_name} → {role_file}")

    # Append to master file (Excel)
    if all_results:
        df_all = pd.DataFrame(all_results)
        if os.path.exists(master_file):
            existing = pd.read_excel(master_file)
            combined = pd.concat([existing, df_all], ignore_index=True)
            with pd.ExcelWriter(master_file, engine="openpyxl") as writer:
                combined.to_excel(writer, index=False)
        else:
            with pd.ExcelWriter(master_file, engine="openpyxl") as writer:
                df_all.to_excel(writer, index=False)

        apply_excel_styling(master_file)
        print(f"Appended results to master file → {master_file}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Evaluate resumes against job descriptions using Gemini.")
    parser.add_argument("--date", required=True, help="Date of resumes in YYYY-MM-DD format")
    args = parser.parse_args()

    genai.configure(api_key=os.getenv("GOOGLE_API_KEY"))

    try:
        datetime.strptime(args.date, "%Y-%m-%d")  # validate format
        evaluate_resumes(args.date)
    except ValueError:
        print("Invalid date format. Please use YYYY-MM-DD.")
